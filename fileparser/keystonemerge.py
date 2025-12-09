import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# 读取两个Excel文件
apis_df = pd.read_excel('keystone_apis.xlsx')
policies_df = pd.read_excel('keystone_policies.xlsx')

# 准备结果列表
matched_results = []
unmatched_apis = []
matched_policy_indices = set()

# 遍历APIs文件中的每一行
for idx, api_row in apis_df.iterrows():
    api_name = api_row['API名称']
    http_method = api_row['HTTP方法']
    url_path = api_row['URL路径']
    endpoint = f"{http_method} {url_path}"
    
    # 在policies文件中查找匹配的Operation
    matched = False
    for policy_idx, policy_row in policies_df.iterrows():
        operation = str(policy_row['Operations']).strip()
        
        # 检查是否完全匹配
        if operation == endpoint:
            matched = True
            matched_policy_indices.add(policy_idx)
            
            # 添加匹配结果
            matched_results.append({
                'API名称': api_name,
                'HTTP方法': http_method,
                'URL路径': url_path,
                'Name': policy_row['策略名称'],
                'Default': policy_row['Default'],
                'Operation': policy_row['Operations'],
                'Scope Types': policy_row['Scope Types'],
                '描述': policy_row['描述']
            })
    
    # 如果没有匹配到，记录该API
    if not matched:
        unmatched_apis.append({
            'API名称': api_name,
            'HTTP方法': http_method,
            'URL路径': url_path,
            'Name': '',
            'Default': '',
            'Operation': '',
            'Scope Types': '',
            '描述': '未匹配到对应的策略'
        })

# 合并匹配和未匹配的结果
all_results = matched_results + unmatched_apis

# 找出未被匹配的policies
unmatched_policies = []
for policy_idx, policy_row in policies_df.iterrows():
    if policy_idx not in matched_policy_indices:
        unmatched_policies.append({
            '策略名称': policy_row['策略名称'],
            'Default': policy_row['Default'],
            'Operations': policy_row['Operations'],
            'Scope Types': policy_row['Scope Types'],
            '描述': policy_row['描述']
        })

# 创建Excel工作簿
wb = Workbook()
wb.remove(wb.active)

# 创建第一个工作表：匹配结果
ws1 = wb.create_sheet('匹配结果')
ws1.append(['API名称', 'HTTP方法', 'URL路径', 'Name', 'Default', 'Operation', 'Scope Types', '描述'])

# 写入数据并合并相同API名称的单元格
if all_results:
    current_row = 2
    i = 0
    while i < len(all_results):
        api_name = all_results[i]['API名称']
        start_row = current_row
        
        # 找出所有相同API名称的行
        while i < len(all_results) and all_results[i]['API名称'] == api_name:
            row_data = [
                all_results[i]['API名称'],
                all_results[i]['HTTP方法'],
                all_results[i]['URL路径'],
                all_results[i]['Name'],
                all_results[i]['Default'],
                all_results[i]['Operation'],
                all_results[i]['Scope Types'],
                all_results[i]['描述']
            ]
            ws1.append(row_data)
            current_row += 1
            i += 1
        
        # 合并API名称、HTTP方法和URL路径列
        if current_row - start_row > 1:
            ws1.merge_cells(f'A{start_row}:A{current_row-1}')
            ws1.merge_cells(f'B{start_row}:B{current_row-1}')
            ws1.merge_cells(f'C{start_row}:C{current_row-1}')
            
            # 设置垂直居中对齐
            for col in ['A', 'B', 'C']:
                ws1[f'{col}{start_row}'].alignment = Alignment(vertical='center', horizontal='left')

# 调整列宽
ws1.column_dimensions['A'].width = 40
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 25
ws1.column_dimensions['D'].width = 30
ws1.column_dimensions['E'].width = 40
ws1.column_dimensions['F'].width = 50
ws1.column_dimensions['G'].width = 20
ws1.column_dimensions['H'].width = 40

# 创建第二个工作表：未匹配的策略
ws2 = wb.create_sheet('未匹配的策略')
ws2.append(['策略名称', 'Default', 'Operations', 'Scope Types', '描述'])

for policy in unmatched_policies:
    ws2.append([
        policy['策略名称'],
        policy['Default'],
        policy['Operations'],
        policy['Scope Types'],
        policy['描述']
    ])

# 调整第二个工作表的列宽
ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 40
ws2.column_dimensions['C'].width = 50
ws2.column_dimensions['D'].width = 20
ws2.column_dimensions['E'].width = 40

# 保存文件
output_filename = 'keystone_merged_result.xlsx'
wb.save(output_filename)

print(f"处理完成！")
print(f"- 总共处理了 {len(apis_df)} 个API")
print(f"- 匹配到 {len(matched_results)} 条记录")
print(f"- 未匹配的API: {len(unmatched_apis)} 个")
print(f"- 未被使用的策略: {len(unmatched_policies)} 条")
print(f"- 结果已保存到: {output_filename}")