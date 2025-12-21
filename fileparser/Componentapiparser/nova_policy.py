import requests
from bs4 import BeautifulSoup
import csv
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def extract_policy_info(url):
    """
    从OpenStack NOVA policy文档中提取策略信息
    """
    # 获取网页内容
    response = requests.get(url)
    response.raise_for_status()
    
    # 解析HTML
    soup = BeautifulSoup(response.content, 'html.parser')
    
    policies = []
    
    # 查找所有的dt标签（包含策略名称）
    dt_elements = soup.find_all('dt')
    
    for dt in dt_elements:
        # 提取策略名称
        code_elem = dt.find('code')
        if not code_elem:
            continue
            
        policy_name = code_elem.get_text(strip=True)
        
        # 获取紧随其后的dd标签（包含详细信息）
        dd = dt.find_next_sibling('dd')
        if not dd:
            continue
        
        # 初始化变量
        default_value = ''
        operations = []
        description = ''
        
        # 提取Default值
        default_field = dd.find('th', string='Default:')
        if default_field:
            default_td = default_field.find_next_sibling('td')
            if default_td:
                default_code = default_td.find('code')
                if default_code:
                    default_value = default_code.get_text(strip=True)
        
        # 提取Operations
        operations_field = dd.find('th', string='Operations:')
        if operations_field:
            operations_td = operations_field.find_next_sibling('td')
            if operations_td:
                li_elements = operations_td.find_all('li')
                for li in li_elements:
                    # 提取HTTP方法
                    strong = li.find('strong')
                    method = strong.get_text(strip=True) if strong else ''
                    
                    # 提取路径
                    code = li.find('code')
                    path = code.get_text(strip=True) if code else ''
                    
                    if method and path:
                        operations.append(f"{method} {path}")
        
        # 提取描述（dd标签中的最后一个p标签）
        p_elements = dd.find_all('p', class_='last')
        if p_elements:
            description = p_elements[-1].get_text(strip=True)
        
        # 如果没有operations，至少添加一条记录
        if not operations:
            operations = ['']
        
        # 为每个operation创建一条记录
        for operation in operations:
            policies.append({
                'name': policy_name,
                'default': default_value,
                'operation': operation,
                'description': description
            })
    
    return policies

def save_to_excel(policies, filename='nova_policies.xlsx'):
    """保存到Excel文件（带格式，多个operations分行显示）"""
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "NOVA Policies"
    
    # 定义表头
    headers = ['Name', 'Default', 'Operation', 'Description']
    
    # 设置表头样式
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    # 定义边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 按策略名称分组，用于合并单元格
    current_row = 2
    prev_name = None
    merge_start_row = 2
    
    # 写入数据
    for i, policy in enumerate(policies):
        ws.cell(row=current_row, column=1).value = policy['name']
        ws.cell(row=current_row, column=2).value = policy['default']
        ws.cell(row=current_row, column=3).value = policy['operation']
        ws.cell(row=current_row, column=4).value = policy['description']
        
        # 设置对齐方式和边框
        for col_num in range(1, 5):
            cell = ws.cell(row=current_row, column=col_num)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
        
        # 检查是否需要合并单元格（同一策略的多个operations）
        if i + 1 < len(policies) and policies[i + 1]['name'] == policy['name']:
            # 下一行是同一个策略，继续
            pass
        else:
            # 当前策略的最后一行，合并name、default和description列
            if current_row > merge_start_row:
                ws.merge_cells(f'A{merge_start_row}:A{current_row}')
                ws.merge_cells(f'B{merge_start_row}:B{current_row}')
                ws.merge_cells(f'D{merge_start_row}:D{current_row}')
                
                # 设置合并单元格的对齐方式
                ws.cell(row=merge_start_row, column=1).alignment = Alignment(
                    vertical='center', horizontal='left', wrap_text=True
                )
                ws.cell(row=merge_start_row, column=2).alignment = Alignment(
                    vertical='center', horizontal='left', wrap_text=True
                )
                ws.cell(row=merge_start_row, column=4).alignment = Alignment(
                    vertical='center', horizontal='left', wrap_text=True
                )
            
            merge_start_row = current_row + 1
        
        current_row += 1
    
    # 调整列宽
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 50
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存文件
    wb.save(filename)
    print(f"数据已保存到 {filename}")

def save_to_csv(policies, filename='nova_policies.csv'):
    """保存到CSV文件"""
    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['name', 'default', 'operation', 'description']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        writer.writeheader()
        for policy in policies:
            writer.writerow(policy)
    
    print(f"数据已保存到 {filename}")

def save_to_json(policies, filename='nova_policies.json'):
    """保存到JSON文件（保持原始结构，operations为列表）"""
    # 重新组织数据结构
    organized_policies = {}
    for policy in policies:
        name = policy['name']
        if name not in organized_policies:
            organized_policies[name] = {
                'name': name,
                'default': policy['default'],
                'operations': [],
                'description': policy['description']
            }
        if policy['operation']:
            organized_policies[name]['operations'].append(policy['operation'])
    
    with open(filename, 'w', encoding='utf-8') as jsonfile:
        json.dump(list(organized_policies.values()), jsonfile, indent=2, ensure_ascii=False)
    
    print(f"数据已保存到 {filename}")

if __name__ == '__main__':
    url = 'https://docs.openstack.org/nova/rocky/configuration/policy.html'
    
    print(f"正在从 {url} 提取数据...")
    policies = extract_policy_info(url)
    
    print(f"共提取到 {len(policies)} 条记录")
    
    # 统计唯一策略数量
    unique_policies = len(set(p['name'] for p in policies))
    print(f"共 {unique_policies} 个唯一策略")
    
    # 保存到Excel（多个operations分行显示，并合并单元格）
    save_to_excel(policies)
    
    # 保存到CSV
    save_to_csv(policies)
    
    # 保存到JSON
    save_to_json(policies)
    
    # 打印前3个策略的示例
    print("\n前3个策略示例：")
    prev_name = None
    count = 0
    for policy in policies:
        if policy['name'] != prev_name:
            count += 1
            if count > 3:
                break
            print(f"\n{count}. {policy['name']}")
            print(f"   Default: {policy['default']}")
            print(f"   Description: {policy['description']}")
            print(f"   Operations:")
            prev_name = policy['name']
        
        if count <= 3:
            print(f"      - {policy['operation']}")