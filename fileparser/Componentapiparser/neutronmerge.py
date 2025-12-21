import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re


def normalize_endpoint(endpoint):
    """
    标准化端点URL，用于匹配
    1. 移除所有空格
    2. 移除 /v3/{project_id}/ 前缀
    """
    # 移除所有空格
    normalized = re.sub(r'\s+', '', str(endpoint).strip())


    return re.sub(r'/v\d+\.\d+/', '/', normalized)

def match_api_with_policy(api_file, policy_file, output_file):
    """
    匹配 API 和 Policy 文件，生成新的 Excel 文件
    
    Args:
        api_file: glance_apis.xlsx 文件路径
        policy_file: glance_policies.xlsx 文件路径
        output_file: 输出文件路径
    """
    # 读取两个 Excel 文件
    print("正在读取文件...")
    df_api = pd.read_excel(api_file)
    df_policy = pd.read_excel(policy_file)
    
    print(f"API 文件共有 {len(df_api)} 条记录")
    print(f"Policy 文件共有 {len(df_policy)} 条记录")
    print("-" * 80)
    
    # 存储匹配结果
    matched_results = []
    matched_policy_indices = set()  # 记录已匹配的 policy 索引
    
    # 遍历 API 列表
    for idx, api_row in df_api.iterrows():
        api_name = api_row['API名称']
        http_method = api_row['HTTP方法']
        endpoint_url = api_row['端点URL']
        full_endpoint = f"{http_method} {endpoint_url}"
        
        print(f"处理 API: {api_name}")
        print(f"  原始端点: {full_endpoint}")
        
        # 标准化 API 端点（移除空格）
        api_operation_normalized = normalize_endpoint(full_endpoint)
        print(f"  标准化后: {api_operation_normalized}")
        
        # 在 policy 中查找匹配的 Operation
        matched = False
        for policy_idx, policy_row in df_policy.iterrows():
            policy_operation = str(policy_row['Operations'])
            
            # 跳过 N/A
            if policy_operation == 'N/A':
                continue
            
            policy_operation_normalized = normalize_endpoint(policy_operation)
            
            # 完全匹配（忽略空格）
            if policy_operation_normalized == api_operation_normalized:
                matched = True
                matched_policy_indices.add(policy_idx)
                
                matched_results.append({
                    'API名称': api_name,
                    'HTTP方法': http_method,
                    'URL路径': endpoint_url,
                    'Policy名称': policy_row['策略名称'],
                    'Default': policy_row['Default'],
                    'Operation': policy_row['Operations'],
                    'Scope Types': policy_row['Scope Types'],
                    'Description': policy_row['描述']
                })
                
                print(f"  ✓ 匹配到: {policy_row['策略名称']}")
                print(f"    Policy Operation: {policy_operation}")
        
        # 如果没有匹配到，也要记录
        if not matched:
            matched_results.append({
                'API名称': api_name,
                'HTTP方法': http_method,
                'URL路径': endpoint_url,
                'Policy名称': 'N/A',
                'Default': 'N/A',
                'Operation': 'N/A',
                'Scope Types': 'N/A',
                'Description': '未匹配到对应的 Policy'
            })
            print(f"  ✗ 未匹配到任何 Policy")
        
        print()
    
    # 查找未被匹配的 policy
    unmatched_policies = []
    for policy_idx, policy_row in df_policy.iterrows():
        if policy_idx not in matched_policy_indices:
            unmatched_policies.append({
                'Policy名称': policy_row['策略名称'],
                'Default': policy_row['Default'],
                'Operation': policy_row['Operations'],
                'Scope Types': policy_row['Scope Types'],
                'Description': policy_row['描述']
            })
    
    # 创建 DataFrame
    df_matched = pd.DataFrame(matched_results)
    df_unmatched = pd.DataFrame(unmatched_policies)
    
    print("-" * 80)
    print(f"匹配完成:")
    print(f"  总共匹配了 {len([r for r in matched_results if r['Policy名称'] != 'N/A'])} 条记录")
    print(f"  未匹配的 API: {len([r for r in matched_results if r['Policy名称'] == 'N/A'])} 条")
    print(f"  未被使用的 Policy: {len(unmatched_policies)} 条")
    
    # 创建 Excel 文件并写入数据（带单元格合并）
    create_excel_with_merge(df_matched, df_unmatched, output_file)
    
    print(f"\n结果已保存到 {output_file}")

def create_excel_with_merge(df_matched, df_unmatched, output_file):
    """
    创建 Excel 文件并合并相同 API 名称的单元格
    """
    wb = Workbook()
    
    # 第一个工作表：匹配结果
    ws_matched = wb.active
    ws_matched.title = "匹配结果"
    
    # 写入表头
    headers = ['API名称', 'HTTP方法', 'URL路径', 'Policy名称', 'Default', 'Operation', 'Scope Types', 'Description']
    ws_matched.append(headers)
    
    # 写入数据并记录需要合并的单元格
    merge_ranges = {}  # {api_name: [(start_row, end_row)]}
    current_api = None
    start_row = 2  # 数据从第2行开始（第1行是表头）
    
    for idx, row in df_matched.iterrows():
        row_num = idx + 2  # Excel 行号（从1开始，第1行是表头）
        
        api_name = row['API名称']
        
        # 检查是否是新的 API
        if api_name != current_api:
            # 如果之前有 API，记录合并范围
            if current_api is not None and row_num - 1 > start_row:
                if current_api not in merge_ranges:
                    merge_ranges[current_api] = []
                merge_ranges[current_api].append((start_row, row_num - 1))
            
            # 更新当前 API
            current_api = api_name
            start_row = row_num
        
        # 写入行数据
        ws_matched.append([
            row['API名称'],
            row['HTTP方法'],
            row['URL路径'],
            row['Policy名称'],
            row['Default'],
            row['Operation'],
            row['Scope Types'],
            row['Description']
        ])
    
    # 处理最后一个 API 的合并
    if current_api is not None and len(df_matched) + 1 > start_row:
        if current_api not in merge_ranges:
            merge_ranges[current_api] = []
        merge_ranges[current_api].append((start_row, len(df_matched) + 1))
    
    # 执行单元格合并（只合并 API名称、HTTP方法、URL路径 列）
    for api_name, ranges in merge_ranges.items():
        for start_row, end_row in ranges:
            if end_row > start_row:  # 只有多行时才合并
                # 合并 API名称（第1列）
                ws_matched.merge_cells(f'A{start_row}:A{end_row}')
                ws_matched[f'A{start_row}'].alignment = Alignment(vertical='center', horizontal='left')
                
                # 合并 HTTP方法（第2列）
                ws_matched.merge_cells(f'B{start_row}:B{end_row}')
                ws_matched[f'B{start_row}'].alignment = Alignment(vertical='center', horizontal='center')
                
                # 合并 URL路径（第3列）
                ws_matched.merge_cells(f'C{start_row}:C{end_row}')
                ws_matched[f'C{start_row}'].alignment = Alignment(vertical='center', horizontal='left')
    
    # 调整列宽
    ws_matched.column_dimensions['A'].width = 30
    ws_matched.column_dimensions['B'].width = 12
    ws_matched.column_dimensions['C'].width = 35
    ws_matched.column_dimensions['D'].width = 30
    ws_matched.column_dimensions['E'].width = 50
    ws_matched.column_dimensions['F'].width = 35
    ws_matched.column_dimensions['G'].width = 20
    ws_matched.column_dimensions['H'].width = 40
    
    # 第二个工作表：未匹配的 Policy
    ws_unmatched = wb.create_sheet("未匹配的Policy")
    
    # 写入未匹配的 Policy 数据
    for r in dataframe_to_rows(df_unmatched, index=False, header=True):
        ws_unmatched.append(r)
    
    # 调整列宽
    ws_unmatched.column_dimensions['A'].width = 30
    ws_unmatched.column_dimensions['B'].width = 50
    ws_unmatched.column_dimensions['C'].width = 35
    ws_unmatched.column_dimensions['D'].width = 20
    ws_unmatched.column_dimensions['E'].width = 40
    
    # 保存文件
    wb.save(output_file)

def main():
    api_file = 'neutron_apis.xlsx'
    policy_file = 'neutron_policies.xlsx'
    output_file = 'neutron_matched.xlsx'
    
    try:
        match_api_with_policy(api_file, policy_file, output_file)
    except Exception as e:
        print(f"\n处理失败: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()