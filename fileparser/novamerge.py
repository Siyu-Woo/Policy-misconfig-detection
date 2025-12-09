import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def fix_merged_cells(df):
    """修正合并单元格造成的空值"""
    # 向下填充空值
    df['Name'] = df['Name'].fillna(method='ffill')
    df['Default'] = df['Default'].fillna(method='ffill')
    df['Description'] = df['Description'].fillna(method='ffill')
    return df

def normalize_endpoint(api_name, http_method, url_path):
    """
    对于带括号且包含action的API名称，修改endpoint格式
    例如: Unlock Server (unlock Action) -> POST /servers/{server_id}/action (unlock)
    """
    # 检查API名称是否包含括号和action
    if '(' in api_name and 'action' in api_name.lower():
        # 提取括号内的内容
        match = re.search(r'\(([^)]+)\s+[Aa]ction\)', api_name)
        if match:
            action_name = match.group(1).strip()
            # 构造新的endpoint格式
            return f"{http_method} {url_path} ({action_name})"
    
    # 默认返回标准格式
    return f"{http_method} {url_path}"

def match_operations(apis_df, policies_df):
    """匹配API和Policy的Operation，支持一对多匹配"""
    matched_records = []
    unmatched_apis = []
    matched_policy_indices = set()
    
    for idx, api_row in apis_df.iterrows():
        api_name = api_row['API名称']
        http_method = api_row['HTTP方法']
        url_path = api_row['URL路径']
        endpoint = api_row['Endpoint']
        
        # 生成标准化的endpoint
        normalized_endpoint = normalize_endpoint(api_name, http_method, url_path)
        
        # 在policies中查找所有匹配的Operation
        api_matches = []
        for policy_idx, policy_row in policies_df.iterrows():
            operation = str(policy_row['Operation']).strip()
            
            if pd.isna(operation) or operation == '' or operation == 'nan':
                continue
            
            # 标准化operation进行比较（去除空格）
            normalized_operation = re.sub(r'\s+', '', operation)
            normalized_endpoint_compare = re.sub(r'\s+', '', normalized_endpoint)
            
            if normalized_operation == normalized_endpoint_compare:
                matched_policy_indices.add(policy_idx)
                
                api_matches.append({
                    'API名称': api_name,
                    'HTTP方法': http_method,
                    'URL路径': url_path,
                    'Name': policy_row['Name'],
                    'Default': policy_row['Default'],
                    'Operation': operation,
                    'Description': policy_row['Description']
                })
        
        if api_matches:
            # 添加所有匹配的记录
            matched_records.extend(api_matches)
        else:
            # 记录未匹配的API
            unmatched_apis.append({
                'API名称': api_name,
                'HTTP方法': http_method,
                'URL路径': url_path,
                'Name': '',
                'Default': '',
                'Operation': '未匹配',
                'Description': ''
            })
    
    # 找出未被匹配的policies
    unmatched_policies = []
    for policy_idx, policy_row in policies_df.iterrows():
        operation = str(policy_row['Operation']).strip()
        
        # 只记录有Operation的policy
        if policy_idx not in matched_policy_indices and operation and operation != 'nan' and operation != '':
            unmatched_policies.append({
                'Name': policy_row['Name'],
                'Default': policy_row['Default'],
                'Operation': operation,
                'Description': policy_row['Description']
            })
    
    return matched_records, unmatched_apis, unmatched_policies

def merge_cells_for_same_api(worksheet, df):
    """合并相同API名称的单元格"""
    if len(df) == 0:
        return
    
    # 定义边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 从第2行开始（第1行是标题）
    start_row = 2
    current_api = None
    merge_start = start_row
    
    for idx, row in df.iterrows():
        current_row = start_row + idx
        api_name = row['API名称']
        
        if current_api is None:
            current_api = api_name
            merge_start = current_row
        elif current_api != api_name:
            # 当API名称改变时，合并前面相同的单元格
            if current_row - 1 > merge_start:
                # 合并API名称、HTTP方法、URL路径列（前3列）
                worksheet.merge_cells(start_row=merge_start, start_column=1, 
                                    end_row=current_row-1, end_column=1)
                worksheet.merge_cells(start_row=merge_start, start_column=2, 
                                    end_row=current_row-1, end_column=2)
                worksheet.merge_cells(start_row=merge_start, start_column=3, 
                                    end_row=current_row-1, end_column=3)
                
                # 设置合并单元格的垂直居中
                for col in range(1, 4):
                    cell = worksheet.cell(row=merge_start, column=col)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            current_api = api_name
            merge_start = current_row
    
    # 处理最后一组
    if len(df) > 0:
        last_row = start_row + len(df) - 1
        if last_row > merge_start:
            worksheet.merge_cells(start_row=merge_start, start_column=1, 
                                end_row=last_row, end_column=1)
            worksheet.merge_cells(start_row=merge_start, start_column=2, 
                                end_row=last_row, end_column=2)
            worksheet.merge_cells(start_row=merge_start, start_column=3, 
                                end_row=last_row, end_column=3)
            
            for col in range(1, 4):
                cell = worksheet.cell(row=merge_start, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # 为所有单元格添加边框
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                   min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border

def save_to_excel(matched_records, unmatched_apis, unmatched_policies, output_file):
    """保存结果到Excel文件，并合并相同API的单元格"""
    # 创建workbook
    workbook = Workbook()
    workbook.remove(workbook.active)  # 删除默认sheet
    
    # 合并匹配和未匹配的API，保持原始顺序
    all_records = matched_records + unmatched_apis
    matched_df = pd.DataFrame(all_records)
    
    # 创建匹配结果工作表
    ws_matched = workbook.create_sheet('匹配结果', 0)
    
    # 写入数据
    for r_idx, row in enumerate(dataframe_to_rows(matched_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_matched.cell(row=r_idx, column=c_idx, value=value)
            
            # 设置标题行样式
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 合并相同API的单元格
    if len(matched_df) > 0:
        merge_cells_for_same_api(ws_matched, matched_df)
    
    # 自动调整列宽
    for column in ws_matched.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws_matched.column_dimensions[column_letter].width = adjusted_width
    
    # 创建未匹配的policies工作表
    if unmatched_policies:
        ws_unmatched = workbook.create_sheet('未匹配的Policies', 1)
        unmatched_policies_df = pd.DataFrame(unmatched_policies)
        
        for r_idx, row in enumerate(dataframe_to_rows(unmatched_policies_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_unmatched.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 自动调整列宽
        for column in ws_unmatched.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws_unmatched.column_dimensions[column_letter].width = adjusted_width
    
    # 保存文件
    workbook.save(output_file)

def main():
    # 读取Excel文件
    print("正在读取nova_apis.xlsx...")
    apis_df = pd.read_excel('nova_apis.xlsx')
    
    print("正在读取nova_policies.xlsx...")
    policies_df = pd.read_excel('nova_policies.xlsx')
    
    # 修正合并单元格
    print("正在修正合并单元格...")
    policies_df = fix_merged_cells(policies_df)
    
    # 匹配操作
    print("正在匹配API和Policy...")
    matched_records, unmatched_apis, unmatched_policies = match_operations(apis_df, policies_df)
    
    # 保存结果
    output_file = 'nova_matched_results.xlsx'
    print(f"正在保存结果到 {output_file}...")
    save_to_excel(matched_records, unmatched_apis, unmatched_policies, output_file)
    
    # 输出统计信息
    print("\n处理完成！")
    print(f"匹配成功的记录数: {len(matched_records)}")
    print(f"未匹配的API: {len(unmatched_apis)}")
    print(f"未匹配的Policies: {len(unmatched_policies)}")
    
    # 统计有多少个API匹配到了多个Policy
    if matched_records:
        api_count = {}
        for record in matched_records:
            api_name = record['API名称']
            api_count[api_name] = api_count.get(api_name, 0) + 1
        
        multi_match = {k: v for k, v in api_count.items() if v > 1}
        if multi_match:
            print(f"\n匹配到多个Policy的API数量: {len(multi_match)}")
            for api_name, count in multi_match.items():
                print(f"  - {api_name}: {count}个匹配")
    
    print(f"\n结果已保存到: {output_file}")

if __name__ == "__main__":
    main()